"""
Affiliate Income Worksheet — backend.

Upload a QuickBooks **Consolidated P&L** export (.xlsx / .xlsm / .csv) and download
the **Affiliate Income Worksheet**: one row per affiliate broker with the broker's
year-to-date Net Income, Y&S's ownership %, Y&S's share, and a column for the new
journal-entry difference.

What it does
------------
1. Finds the affiliate **entity header row** (e.g. "YSM Tickets", "YSS Tickets", …)
   and the **"Net Income"** row in the P&L.
2. For each configured broker, reads that entity column's Net Income into
   **Total YTD Net Inc thru Last Month** (column B).
3. Fills in the fixed **Y&S % Ownership** (column C) and writes the live formulas
   **Y&S Share = B*C** (column D) and **Difference = D-E** (column F).
4. **Column E** ("…already in the P&L") is read from the **K-1 income** lines booked
   on the YS Affiliates LLC entity. That K-1 income lags one month (YTD through two
   months ago), so column F (= D - E) is the incremental journal entry for last month.
5. **Date** (column G) is the month-end of the last month in the P&L period
   ("January-June, 2026" -> 6/30/2026). Auto-detected, overridable in the UI.

The output workbook has two tabs: **Journal Entries** (the worksheet above) and
**Consolidated P&L** (the uploaded report, embedded for reference).

The broker list, each broker's P&L column name, and ownership % live in BROKERS below
so they are easy to maintain. Columns are matched to the P&L by (normalized) header
name, so column order in the export does not matter.
"""

import io
import os
import csv
import re
import time
import uuid
import shutil
import calendar
import tempfile
import datetime as dt
from copy import copy as _copy

from flask import Flask, request, jsonify, send_file, send_from_directory, abort
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__, static_folder=None)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STORE_DIR = os.path.join(tempfile.gettempdir(), "affinc_store")
os.makedirs(STORE_DIR, exist_ok=True)


# =========================================================================== #
# BROKER CONFIG — edit to add/remove affiliates or change ownership %.
#   (Display name, P&L entity column header, Y&S ownership fraction,
#    K-1 row label, Inv account label)
# - The entity column header feeds column B (Total YTD Net Inc thru Last Month)
#   on the P&L; it is also the affiliate's column on the Balance Sheet.
# - The K-1 row label feeds column E (Y&S Share of Net Inc already in the P&L):
#   the "K-1 - …" income line booked on the YS Affiliates LLC entity. That K-1
#   income lags one month (it is YTD through two months ago, not last month), so
#   column F (= D - E) is the incremental journal entry for the latest month.
# - The Inv account label is the "Inv - …" investment account on YS Affiliates
#   LLC's books, used by the Balance Sheet reconciliation.
# All are matched case/space-insensitively, so column/row order does not matter.
# A broker whose column or row is absent is still written (blank) and warned.
# =========================================================================== #
BROKERS = [
    ("YSM",         "YSM Tickets",         0.50, "K-1 - YSM (Grossman)",  "Inv - YSM (Grossman)"),
    ("YSS",         "YSS Tickets",         0.50, "K-1 - YSS (Sternbuch)", "Inv - YSS (Sternbuch)"),
    ("YSP",         "YSP Tickets",         0.50, "K-1 - YSP (Pollak)",    "Inv - YSP (Pollak)"),
    ("YS Levine",   "YS Levine Tickets",   0.50, "K-1 - YS Levine",       "Inv - YS Levine"),
    ("YS Levovitz", "YS Levovitz Tickets", 0.50, "K-1 - YS Levovitz",     "Inv - YS Levovitz"),
    ("YS Chase",    "YS Chase Tickets",    0.50, "K-1 - YS Chase",        "Inv - YS Chase"),
    ("YS Asher",    "YS Asher Tickets",    0.50, "K-1 - YS Asher",        "Inv - YS Asher"),
    ("YS Katz",     "YS Katz Tickets",     0.50, "K-1 - YS Katz",         "Inv - YS Katz"),
    ("YSKG",        "YSKG Tickets",        0.25, "K-1 - YSKG",            "Inv - YSKG"),
    ("YSTL",        "YS TL Tickets",       0.35, "K-1 - YS TL",           "Inv - YS TL"),
    ("YS Waxler",   "YSW Tickets",         0.50, "K-1 - YSW (Waxler)",    "Inv - YSW (Waxler)"),
]

# Entity column the K-1 income and the Inv accounts are booked on (YS Affiliates
# LLC). Column E source on the P&L; the Inv balances on the Balance Sheet.
K1_SOURCE_COLUMN = "YS Affiliates"

# Balance Sheet row labels used by the reconciliation, read from each affiliate's
# own column. "Total for Equity - Y&S" is Y&S's capital account on the affiliate's
# books and is already net of contributions less distributions.
EQUITY_YS_LABEL = "Total for Equity - Y&S"
RETAINED_EARNINGS_LABEL = "Retained Earnings"

# Entity columns expected on the P&L that are intentionally NOT broker rows
# (parent / roll-up / report columns). Used only to decide whether an unmapped
# "… Tickets" column is worth flagging as a possible new affiliate.
KNOWN_NON_BROKERS = {
    "y&s tickets", "ys needle tickets", "ys affiliates",
    "eliminations", "total",
}

NET_INCOME_LABEL = "net income"      # exact (normalized) row label to read


# =========================================================================== #
# Cell / file helpers
# =========================================================================== #

_NUM_FORMULA = re.compile(r"-?\d+(\.\d+)?")


def _cleanup_old(max_age_seconds=12 * 3600):
    now = time.time()
    for name in os.listdir(STORE_DIR):
        path = os.path.join(STORE_DIR, name)
        try:
            if os.path.isdir(path) and now - os.path.getmtime(path) > max_age_seconds:
                shutil.rmtree(path, ignore_errors=True)
        except OSError:
            pass


def _amount(v):
    """Coerce a cell to float. Handles numbers, '=123.45' literal formulas,
    '$1,234.56', and '(123)' negatives. Cell-ref formulas (=B6+C6) -> None."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return None
    if s.startswith("="):
        body = s[1:]
        return float(body) if _NUM_FORMULA.fullmatch(body) else None
    neg = s.startswith("(") and s.endswith(")")
    s = re.sub(r"[^0-9.\-]", "", s)
    if s in ("", "-", "."):
        return None
    try:
        f = float(s)
        return -f if neg else f
    except ValueError:
        return None


def _norm(s):
    return re.sub(r"\s+", " ", str(s).strip().lower()) if s is not None else ""


def _first_text(row):
    for c in row:
        if isinstance(c, str) and c.strip():
            return c.strip()
    return None


def _rows_from_upload(filename, data):
    """Return a list-of-lists for the uploaded report. For workbooks the largest
    sheet is used and cached formula values are read (data_only=True), since QBO
    exports carry literal values."""
    low = filename.lower()
    if low.endswith(".csv"):
        text = data.decode("utf-8-sig", errors="replace")
        return [list(r) for r in csv.reader(io.StringIO(text))]
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    best, best_score = None, -1
    for ws in wb.worksheets:
        score = ws.max_row * (ws.max_column or 1)
        if score > best_score:
            best, best_score = ws, score
    rows = [list(r) for r in best.iter_rows(values_only=True)]
    wb.close()
    return rows


def _safe_filename(s):
    return re.sub(r'[\\/:*?"<>|]+', " ", s).strip() if s else s


# =========================================================================== #
# Period / date detection
# =========================================================================== #

_MONTHS = {m.lower(): i for i, m in enumerate(calendar.month_name) if m}
_MONTHS.update({m.lower(): i for i, m in enumerate(calendar.month_abbr) if m})
_MONTH_RE = re.compile(
    r"\b(" + "|".join(sorted((re.escape(m) for m in _MONTHS), key=len, reverse=True)) + r")\b",
    re.I,
)


def _find_period_end(rows):
    """Month-end of the last month named in the report's period line.
    Scans only the top rows so the print-timestamp footer is ignored.
    'January-June, 2026' -> date(2026, 6, 30); 'May 2026' -> date(2026, 5, 31)."""
    for row in rows[:6]:
        for cell in row:
            if not isinstance(cell, str) or not cell.strip():
                continue
            year = re.search(r"\b(20\d{2})\b", cell)
            months = _MONTH_RE.findall(cell)
            if year and months:
                mon = _MONTHS[months[-1].lower()]
                yr = int(year.group(1))
                last = calendar.monthrange(yr, mon)[1]
                return dt.date(yr, mon, last)
    return None


# =========================================================================== #
# P&L parsing
# =========================================================================== #

def _broker_header_keys():
    return {_norm(h) for _disp, h, _pct, _k1, _inv in BROKERS}


def find_header_row(rows):
    """The row containing the affiliate entity headers — the one matching the most
    configured broker columns. Returns (row_index, {normalized_header: col_index})."""
    want = _broker_header_keys()
    best_i, best_hits = None, 0
    for i, row in enumerate(rows):
        hits = sum(1 for c in row if isinstance(c, str) and _norm(c) in want)
        if hits > best_hits:
            best_i, best_hits = i, hits
    if best_i is None or best_hits < 2:
        return None, {}
    header_map = {}
    for ci, c in enumerate(rows[best_i]):
        if isinstance(c, str) and c.strip():
            header_map.setdefault(_norm(c), ci)
    return best_i, header_map


def find_net_income_row(rows, start=0):
    """Index of the row whose first text cell is exactly 'Net Income' (normalized),
    distinct from 'Net Operating Income' / 'Net Other Income'. Last match wins."""
    found = None
    for i in range(start, len(rows)):
        label = _first_text(rows[i])
        if label is not None and _norm(label) == NET_INCOME_LABEL:
            found = i
    return found


def parse_pl(rows):
    """Return (broker_records, warnings). Each record:
    {Broker, NetIncome (float|None), Ownership (float), Booked (float|None)}.
    NetIncome feeds column B; Booked (the K-1 income already in the P&L) feeds
    column E."""
    hdr_i, header_map = find_header_row(rows)
    if hdr_i is None:
        raise ValueError(
            "Could not find the affiliate entity header row (e.g. 'YSM Tickets', "
            "'YSS Tickets', …) in the P&L. Is this the Consolidated P&L export?")

    ni_i = find_net_income_row(rows, start=hdr_i + 1)
    if ni_i is None:
        raise ValueError("Could not find a 'Net Income' row in the P&L.")
    ni_row = rows[ni_i]

    # Column the K-1 income is booked on (for column E).
    k1_col = header_map.get(_norm(K1_SOURCE_COLUMN))
    if k1_col is None:
        k1_col = header_map.get("total")
    # Map each configured K-1 row label -> its value on that column.
    k1_labels = {_norm(k1): None for _disp, _h, _p, k1, _inv in BROKERS}
    if k1_col is not None:
        for row in rows:
            label = _first_text(row)
            key = _norm(label)
            if key in k1_labels and k1_col < len(row):
                k1_labels[key] = _amount(row[k1_col])

    records, missing, missing_k1 = [], [], []
    for display, header, pct, k1, _inv in BROKERS:
        col = header_map.get(_norm(header))
        val = _amount(ni_row[col]) if (col is not None and col < len(ni_row)) else None
        if col is None:
            missing.append(f"{display} (column \u201c{header}\u201d)")
        booked = k1_labels.get(_norm(k1))
        if booked is None:
            missing_k1.append(f"{display} (row \u201c{k1}\u201d)")
            booked = 0.0
        records.append({"Broker": display, "NetIncome": val,
                        "Ownership": pct, "Booked": booked})

    # Flag any "… Tickets" entity column on the P&L that we don't map and isn't a
    # known parent/roll-up — a hint that a new affiliate may need adding to BROKERS.
    mapped = _broker_header_keys()
    extras = []
    for key in header_map:
        if key in mapped or key in KNOWN_NON_BROKERS:
            continue
        if key.endswith("tickets"):
            extras.append(rows[hdr_i][header_map[key]])

    warnings = []
    if missing:
        warnings.append("These brokers were not found in the P&L and were left blank: "
                        + "; ".join(missing) + ".")
    if k1_col is None:
        warnings.append(f"Could not find the \u201c{K1_SOURCE_COLUMN}\u201d (or Total) "
                        "column, so the K-1 amounts already booked (column E) are 0. "
                        "Verify column E before posting.")
    elif missing_k1:
        warnings.append("No K-1 income row was found for these brokers, so their "
                        "column E was set to 0: " + "; ".join(missing_k1) + ".")
    if extras:
        warnings.append("The P&L has these entity columns that are not set up as "
                        "brokers (ignored): " + ", ".join(map(str, extras))
                        + ". Add them to the broker list if they should be included.")
    return records, warnings


# =========================================================================== #
# Balance Sheet parsing — reconciliation
#
# For each affiliate, the "Inv - …" account on YS Affiliates LLC's books should
# equal, per the affiliate's own column on the Balance Sheet:
#
#     Total for Equity - Y&S  +  Y&S% x (Retained Earnings + Net Income)
#
# "Total for Equity - Y&S" is already net of capital contributions less
# distributions, so it adds directly. Any residual is the variance to chase.
# =========================================================================== #

def _find_row_by_label(rows, label, start=0):
    """Index of the row whose first text cell equals `label` (normalized).
    Last match wins, so a repeated label resolves to the final occurrence."""
    want = _norm(label)
    found = None
    for i in range(start, len(rows)):
        first = _first_text(rows[i])
        if first is not None and _norm(first) == want:
            found = i
    return found


def _cell(rows, row_i, col_i):
    if row_i is None or col_i is None or row_i >= len(rows):
        return None
    row = rows[row_i]
    return _amount(row[col_i]) if col_i < len(row) else None


def parse_balance_sheet(rows):
    """Return (records, warnings, keep_rows). Each record:
    {Broker, InvAccount, InvBalance, EquityYS, RetainedEarnings, NetIncome,
     Ownership}. keep_rows is the sorted set of row indices the reconciliation
     actually reads (report header rows + the Inv / equity / earnings rows), used
     to trim the embedded Balance Sheet tab."""
    hdr_i, header_map = find_header_row(rows)
    if hdr_i is None:
        raise ValueError(
            "Could not find the affiliate entity header row (e.g. 'YSM Tickets', "
            "'YSS Tickets', …) in the Balance Sheet. Is this the Consolidated "
            "Balance Sheet export?")

    aff_col = header_map.get(_norm(K1_SOURCE_COLUMN))
    if aff_col is None:
        raise ValueError(
            f"Could not find the \u201c{K1_SOURCE_COLUMN}\u201d column on the "
            "Balance Sheet — that is where the Inv accounts are booked.")

    eq_i = _find_row_by_label(rows, EQUITY_YS_LABEL, start=hdr_i + 1)
    re_i = _find_row_by_label(rows, RETAINED_EARNINGS_LABEL, start=hdr_i + 1)
    ni_i = find_net_income_row(rows, start=hdr_i + 1)
    if eq_i is None:
        raise ValueError(f"Could not find a \u201c{EQUITY_YS_LABEL}\u201d row "
                         "on the Balance Sheet.")
    if re_i is None:
        raise ValueError(f"Could not find a \u201c{RETAINED_EARNINGS_LABEL}\u201d "
                         "row on the Balance Sheet.")
    if ni_i is None:
        raise ValueError("Could not find a 'Net Income' row on the Balance Sheet.")

    # Report header rows (title, "As of …", entity headers) plus the three
    # shared rows every broker reads.
    keep = set(range(0, hdr_i + 1)) | {re_i, ni_i, eq_i}

    records, missing_col, missing_inv = [], [], []
    for display, header, pct, _k1, inv_label in BROKERS:
        col = header_map.get(_norm(header))
        if col is None:
            missing_col.append(f"{display} (column \u201c{header}\u201d)")

        inv_i = _find_row_by_label(rows, inv_label, start=hdr_i + 1)
        if inv_i is None:
            missing_inv.append(f"{display} (row \u201c{inv_label}\u201d)")
        else:
            keep.add(inv_i)

        records.append({
            "Broker": display,
            "InvAccount": inv_label,
            "InvBalance": _cell(rows, inv_i, aff_col),
            "EquityYS": _cell(rows, eq_i, col),
            "RetainedEarnings": _cell(rows, re_i, col),
            "NetIncome": _cell(rows, ni_i, col),
            "Ownership": pct,
        })

    warnings = []
    if missing_col:
        warnings.append("These affiliates have no column on the Balance Sheet, so "
                        "their equity/earnings were left blank: "
                        + "; ".join(missing_col) + ".")
    if missing_inv:
        warnings.append("These Inv accounts were not found on the Balance Sheet, so "
                        "their balance was left blank: " + "; ".join(missing_inv) + ".")
    return records, warnings, sorted(keep)


# =========================================================================== #
# Workbook builder — matches Affiliate_Income_Worksheet styling exactly
# =========================================================================== #

CUR = '"$"#,##0.00;[Red]-"$"#,##0.00'    # negatives render in red
PCT = '0.00%'
DATEFMT = "mm-dd-yy"
FONT = "Calibri"
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")

COLUMNS = [
    ("Broker", "A", 10.71, "General"),
    ("Total YTD Net Inc\nthru Last Month", "B", 16.14, CUR),
    ("Y&S %\nOwnership", "C", 10.71, PCT),
    ("Y&S Share of Net Inc\nthru Last Month", "D", 19.43, CUR),
    ("Y&S Share of Net Inc\nalready in the P&L", "E", 19.43, CUR),
    ("Difference - New\nJournal Entry", "F", 16.29, CUR),
    ("Date", "G", 14.0, DATEFMT),
]


# --- source-tab embedding ---------------------------------------------------- #

def _sanitize_sheet_title(title):
    t = re.sub(r"[:\\/?*\[\]]", " ", str(title)).strip()
    return (t or "Source")[:31]


def _unique_title(wb, title):
    base, n, out = title, 2, title
    while out in wb.sheetnames:
        out = _sanitize_sheet_title(f"{base} {n}")
        n += 1
    return out


def _largest_sheet(workbook):
    best, best_score = None, -1
    for s in workbook.worksheets:
        score = (s.max_row or 0) * (s.max_column or 1)
        if score > best_score:
            best, best_score = s, score
    return best


def _dump_rows_sheet(wb, title, rows):
    """Fallback: write parsed values into a new sheet (CSV uploads, or if a faithful
    copy fails). Numbers stored as numbers; text left-aligned."""
    ws = wb.create_sheet(title)
    left = Alignment(horizontal="left", vertical="center")
    widths = {}
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(r, c)
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.value = float(val)
                cell.number_format = '#,##0.00'
            else:
                num = _amount(val)
                if num is not None and str(val).strip() not in ("", "-"):
                    cell.value = num
                    cell.number_format = '#,##0.00'
                else:
                    cell.value = val if val not in ("", None) else None
                    cell.alignment = left
            cell.font = Font(name=FONT, size=11)
            widths[c] = max(widths.get(c, 0), len(str(val)) if val is not None else 0)
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = min(max(w + 2, 8), 60)
    return ws


def _copy_source_faithfully(wb, title, data, keep_rows=None):
    """Copy the largest sheet of an uploaded xlsx/xlsm into wb, preserving values,
    number formats, fonts, fills, borders, alignment, widths, heights, merges.
    If keep_rows (0-based source row indices) is given, only those rows are copied
    and they are packed together sequentially."""
    src = load_workbook(io.BytesIO(data), data_only=True)
    best = _largest_sheet(src)
    ws = wb.create_sheet(title)

    keep = None if keep_rows is None else {i + 1 for i in keep_rows}   # -> 1-based
    dest_r, row_map = 0, {}
    for row in best.iter_rows():
        if not row:
            continue
        src_r = row[0].row
        if keep is not None and src_r not in keep:
            continue
        dest_r += 1
        row_map[src_r] = dest_r
        for oc in row:
            nc = ws.cell(row=dest_r, column=oc.column, value=oc.value)
            if oc.has_style:
                nc.font = _copy(oc.font)
                nc.fill = _copy(oc.fill)
                nc.border = _copy(oc.border)
                nc.alignment = _copy(oc.alignment)
                nc.number_format = oc.number_format

    for letter, dim in best.column_dimensions.items():
        if dim.width:
            ws.column_dimensions[letter].width = dim.width
    for idx, dim in best.row_dimensions.items():
        if dim.height and idx in row_map:
            ws.row_dimensions[row_map[idx]].height = dim.height
    if keep is None:      # merges only make sense on an unfiltered copy
        for mc in list(best.merged_cells.ranges):
            ws.merge_cells(str(mc))
    src.close()
    return ws


def _freeze_row_for(rows):
    """Row to freeze above on the source tab: just below the entity-header row so
    the company/entity names stay visible while scrolling. Falls back to row 5
    (freeze rows 1-4)."""
    if rows:
        hdr_i, _ = find_header_row(rows)
        if hdr_i is not None:
            return hdr_i + 2
    return 5


def _append_source_sheet(wb, filename, data, rows_fallback=None,
                         title="Consolidated P&L", keep_rows=None):
    """Add the uploaded report as a tab. xlsx/xlsm copied with formatting; CSV (or
    any failure) falls back to a plain values dump. If keep_rows is given, only
    those source rows are included. Rows through the entity-header row are frozen
    so company names stick. Never raises."""
    title = _unique_title(wb, _sanitize_sheet_title(title))
    ws = None
    if not filename.lower().endswith(".csv"):
        try:
            ws = _copy_source_faithfully(wb, title, data, keep_rows=keep_rows)
        except Exception:
            if title in wb.sheetnames:
                del wb[title]
            ws = None
    if ws is None:
        rows = rows_fallback if rows_fallback is not None else _rows_from_upload(filename, data)
        if keep_rows is not None:
            rows = [rows[i] for i in keep_rows if i < len(rows)]
        ws = _dump_rows_sheet(wb, title, rows)
        rows_fallback = rows
    ws.freeze_panes = f"A{_freeze_row_for(rows_fallback)}"
    return ws


def build_workbook(records, pay_date, source_files=None, source_rows=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Journal Entries"

    # Header row
    ws.append([c[0] for c in COLUMNS])
    ws.row_dimensions[1].height = 30
    for ci, (label, letter, width, _fmt) in enumerate(COLUMNS, start=1):
        cell = ws.cell(1, ci)
        cell.font = Font(name=FONT, size=11, bold=True)
        cell.alignment = CENTER_WRAP
        ws.column_dimensions[letter].width = width

    # Data rows
    for n, rec in enumerate(records):
        r = n + 2
        ws.cell(r, 1, rec["Broker"])
        ni = rec["NetIncome"]
        ws.cell(r, 2, ni if ni is not None else None)        # B  Net Income
        ws.cell(r, 3, rec["Ownership"])                       # C  ownership %
        ws.cell(r, 4, f"=B{r}*C{r}")                          # D  Y&S share
        booked = rec.get("Booked")
        ws.cell(r, 5, booked if booked is not None else None) # E  K-1 already booked
        ws.cell(r, 6, f"=D{r}-E{r}")                          # F  new JE difference
        ws.cell(r, 7, dt.datetime(pay_date.year, pay_date.month, pay_date.day))  # G

        for ci, (_label, _letter, _w, fmt) in enumerate(COLUMNS, start=1):
            cell = ws.cell(r, ci)
            cell.font = Font(name=FONT, size=11)
            cell.alignment = CENTER
            if isinstance(fmt, str):
                cell.number_format = fmt

    # Second tab: the uploaded P&L, embedded for reference.
    for fn, data in (source_files or []):
        _append_source_sheet(wb, fn, data, rows_fallback=source_rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =========================================================================== #
# Reconciliation workbook builder
# =========================================================================== #

RECON_COLUMNS = [
    ("Broker",                     "A", 12.0,  "General"),
    ("Inv Balance\nper Books",     "B", 16.0,  CUR),
    ("Total for\nEquity - Y&S",    "C", 16.0,  CUR),
    ("Retained\nEarnings",         "D", 16.0,  CUR),
    ("Net Income",                 "E", 16.0,  CUR),
    ("Y&S %\nOwnership",           "F", 11.0,  PCT),
    ("Y&S Share of\nRE + Net Inc", "G", 17.0,  CUR),
    ("Expected Inv\nBalance",      "H", 17.0,  CUR),
    ("Difference\n(Actual - Exp)", "I", 17.0,  CUR),
]


def build_recon_workbook(records, as_of, source_files=None, source_rows=None,
                         keep_rows=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    ws.append([c[0] for c in RECON_COLUMNS])
    ws.row_dimensions[1].height = 30
    for ci, (label, letter, width, _fmt) in enumerate(RECON_COLUMNS, start=1):
        cell = ws.cell(1, ci)
        cell.font = Font(name=FONT, size=11, bold=True)
        cell.alignment = CENTER_WRAP
        ws.column_dimensions[letter].width = width

    for n, rec in enumerate(records):
        r = n + 2
        ws.cell(r, 1, rec["Broker"])
        ws.cell(r, 2, rec["InvBalance"])
        ws.cell(r, 3, rec["EquityYS"])
        ws.cell(r, 4, rec["RetainedEarnings"])
        ws.cell(r, 5, rec["NetIncome"])
        ws.cell(r, 6, rec["Ownership"])
        ws.cell(r, 7, f"=F{r}*(D{r}+E{r})")           # Y&S share of RE + NI
        ws.cell(r, 8, f"=C{r}+G{r}")                   # expected Inv balance
        ws.cell(r, 9, f"=B{r}-H{r}")                   # difference

        for ci, (_l, _let, _w, fmt) in enumerate(RECON_COLUMNS, start=1):
            cell = ws.cell(r, ci)
            cell.font = Font(name=FONT, size=11)
            cell.alignment = CENTER
            if isinstance(fmt, str) and fmt != "General":
                cell.number_format = fmt

    ws.freeze_panes = "A2"

    # Second tab: the uploaded Balance Sheet, trimmed to the rows the
    # reconciliation actually reads.
    for fn, data in (source_files or []):
        _append_source_sheet(wb, fn, data, rows_fallback=source_rows,
                             title="Consolidated Balance Sheet",
                             keep_rows=keep_rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =========================================================================== #
# Routes
# =========================================================================== #

@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "index.html")


@app.route("/process", methods=["POST"])
def process():
    pl = [(f.filename, f.read()) for f in request.files.getlist("pl") if f.filename]
    if not pl:
        return jsonify({"error": "Please upload a Consolidated P&L."}), 400

    override = None
    raw = (request.form.get("month_end") or "").strip()
    if raw:
        try:
            override = dt.datetime.strptime(raw, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({"error": "Month-end date must be YYYY-MM-DD."}), 400

    try:
        rows = []
        for fn, data in pl:
            rows.extend(_rows_from_upload(fn, data))
        records, warnings = parse_pl(rows)
        if all(r["NetIncome"] is None for r in records):
            raise ValueError("No affiliate Net Income values were read from the P&L.")

        detected = _find_period_end(rows)
        pay_date = override or detected or dt.date.today()
        if not override and detected is None:
            warnings.insert(0, "Could not detect the period end date from the P&L; "
                               "used today's date. Set the month-end date and rerun.")
    except Exception as exc:
        return jsonify({"error": f"{type(exc).__name__}: {exc}"}), 500

    data_out = build_workbook(records, pay_date, source_files=pl, source_rows=rows)

    token = uuid.uuid4().hex
    fn = f"Affiliate Income Worksheet {pay_date:%B %Y}.xlsx"
    folder = os.path.join(STORE_DIR, token)
    os.makedirs(folder, exist_ok=True)
    with open(os.path.join(folder, fn), "wb") as fh:
        fh.write(data_out)
    _cleanup_old()

    return jsonify({
        "as_of": pay_date.strftime("%B %d, %Y"),
        "brokers": sum(1 for r in records if r["NetIncome"] is not None),
        "filename": fn,
        "download_url": f"/download/{token}",
        "warnings": warnings,
    })


@app.route("/reconcile", methods=["POST"])
def reconcile():
    bs = [(f.filename, f.read())
          for f in request.files.getlist("balance_sheet") if f.filename]
    if not bs:
        return jsonify({"error": "Please upload a Consolidated Balance Sheet."}), 400

    try:
        rows = []
        for fn, data in bs:
            rows.extend(_rows_from_upload(fn, data))
        records, warnings, keep_rows = parse_balance_sheet(rows)
        if all(r["InvBalance"] is None for r in records):
            raise ValueError("No Inv account balances were read from the Balance "
                             "Sheet.")
        as_of = _find_period_end(rows) or dt.date.today()
    except Exception as exc:
        return jsonify({"error": f"{type(exc).__name__}: {exc}"}), 500

    data_out = build_recon_workbook(records, as_of, source_files=bs,
                                    source_rows=rows, keep_rows=keep_rows)

    token = uuid.uuid4().hex
    fn = f"Balance Sheet Reconciliation {as_of:%B %Y}.xlsx"
    folder = os.path.join(STORE_DIR, token)
    os.makedirs(folder, exist_ok=True)
    with open(os.path.join(folder, fn), "wb") as fh:
        fh.write(data_out)
    _cleanup_old()

    return jsonify({
        "as_of": as_of.strftime("%B %d, %Y"),
        "filename": fn,
        "download_url": f"/download/{token}",
        "warnings": warnings,
    })


@app.route("/download/<token>")
def download(token):
    folder = os.path.join(STORE_DIR, os.path.basename(token))
    if not os.path.isdir(folder):
        abort(404)
    xlsx = [f for f in os.listdir(folder) if f.lower().endswith(".xlsx")]
    if not xlsx:
        abort(404)
    pick = xlsx[0]
    return send_file(
        os.path.join(folder, pick),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=pick)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
